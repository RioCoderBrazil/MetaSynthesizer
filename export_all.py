#!/usr/bin/env python3
"""
MetaSynthesizer Export Utility
Export all extraction results to various formats
"""

import argparse
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExportManager:
    """Handle exports to various formats"""
    
    def __init__(self, data_dir: Path = Path("03_extracted_data")):
        self.data_dir = Path(data_dir)
        self.extractions = self._load_all_extractions()
    
    def _load_all_extractions(self) -> List[Dict[str, Any]]:
        """Load all extraction results"""
        extractions = []
        
        # Try to load combined file first
        combined_file = self.data_dir / "all_extractions.json"
        if combined_file.exists():
            with open(combined_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        
        # Otherwise load individual files
        for json_file in self.data_dir.glob("*_extracted.json"):
            if "final_merged" not in json_file.name:  # Skip the merged file
                with open(json_file, 'r', encoding='utf-8') as f:
                    extractions.append(json.load(f))
        
        return extractions
    
    def export_to_excel(self, output_path: Path):
        """Export all data to Excel with multiple sheets"""
        print(f"📊 Exporting to Excel: {output_path}")
        
        wb = openpyxl.Workbook()
        
        # Create overview sheet
        self._create_overview_sheet(wb)
        
        # Create category sheets
        self._create_risk_sheet(wb)
        self._create_recommendations_sheet(wb)
        self._create_findings_sheet(wb)
        self._create_metadata_sheet(wb)
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Save workbook
        wb.save(output_path)
        print(f"✅ Excel export complete: {output_path}")
    
    def export_to_csv(self, output_dir: Path):
        """Export data to multiple CSV files"""
        print(f"📄 Exporting to CSV: {output_dir}")
        output_dir.mkdir(exist_ok=True)
        
        # Export risks
        risks_data = self._extract_all_risks()
        if risks_data:
            risks_df = pd.DataFrame(risks_data)
            risks_df.to_csv(output_dir / "risks.csv", index=False)
        
        # Export recommendations
        recs_data = self._extract_all_recommendations()
        if recs_data:
            recs_df = pd.DataFrame(recs_data)
            recs_df.to_csv(output_dir / "recommendations.csv", index=False)
        
        # Export findings
        findings_data = self._extract_all_findings()
        if findings_data:
            findings_df = pd.DataFrame(findings_data)
            findings_df.to_csv(output_dir / "findings.csv", index=False)
        
        # Export metadata
        metadata = self._extract_metadata()
        if metadata:
            meta_df = pd.DataFrame(metadata)
            meta_df.to_csv(output_dir / "metadata.csv", index=False)
        
        print(f"✅ CSV export complete: {output_dir}")
    
    def export_summary_report(self, output_path: Path):
        """Generate a markdown summary report"""
        print(f"📝 Generating summary report: {output_path}")
        
        report = []
        report.append("# MetaSynthesizer Extraction Summary Report")
        report.append(f"\n**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append(f"\n**Total Documents Processed**: {len(self.extractions)}")
        
        # Overall statistics
        report.append("\n## 📊 Overall Statistics")
        
        completeness_scores = []
        for ext in self.extractions:
            if "meta" in ext and "completeness_score" in ext["meta"]:
                completeness_scores.append(ext["meta"]["completeness_score"])
        
        if completeness_scores:
            avg_completeness = sum(completeness_scores) / len(completeness_scores)
            report.append(f"\n- **Average Completeness**: {avg_completeness:.1f}%")
            report.append(f"- **Highest Completeness**: {max(completeness_scores):.1f}%")
            report.append(f"- **Lowest Completeness**: {min(completeness_scores):.1f}%")
        
        # Risk summary
        report.append("\n## ⚠️ Risk Summary")
        risks = self._extract_all_risks()
        risk_levels = {}
        for risk in risks:
            level = risk.get("risk_level", "unknown")
            risk_levels[level] = risk_levels.get(level, 0) + 1
        
        for level, count in risk_levels.items():
            report.append(f"- **{level.capitalize()} Risk**: {count} documents")
        
        # Recommendations summary
        report.append("\n## 📋 Recommendations Summary")
        recs = self._extract_all_recommendations()
        status_counts = {}
        for rec in recs:
            status = rec.get("status", "unknown")
            status_counts[status] = status_counts.get(status, 0) + 1
        
        report.append(f"- **Total Recommendations**: {len(recs)}")
        for status, count in status_counts.items():
            report.append(f"- **{status.capitalize()}**: {count}")
        
        # Findings summary
        report.append("\n## 🔍 Findings Summary")
        findings = self._extract_all_findings()
        priority_counts = {}
        for finding in findings:
            priority = finding.get("priority", "unknown")
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
        
        report.append(f"- **Total Findings**: {len(findings)}")
        for priority, count in priority_counts.items():
            report.append(f"- **{priority.capitalize()} Priority**: {count}")
        
        # Category coverage
        report.append("\n## 📊 Category Coverage")
        categories = [
            "risk_assessment", "recommendation_tracking", "finding_summary",
            "environmental_impact", "timeline_analysis", "stakeholder_info",
            "compliance_status", "audit_metadata", "financial_data"
        ]
        
        for cat in categories:
            filled_count = sum(
                1 for ext in self.extractions
                if ext.get("categories", {}).get(cat, {})
            )
            coverage = (filled_count / len(self.extractions) * 100) if self.extractions else 0
            report.append(f"- **{cat.replace('_', ' ').title()}**: {coverage:.1f}% coverage")
        
        # Document list
        report.append("\n## 📄 Processed Documents")
        for i, ext in enumerate(self.extractions, 1):
            doc_id = ext.get("document_id", "Unknown")
            completeness = ext.get("meta", {}).get("completeness_score", 0)
            report.append(f"{i}. {doc_id} - {completeness:.1f}% complete")
        
        # Write report
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(report))
        
        print(f"✅ Summary report generated: {output_path}")
    
    def _create_overview_sheet(self, wb):
        """Create overview sheet in Excel"""
        ws = wb.create_sheet("Overview")
        
        # Headers
        headers = ["Document ID", "Completeness %", "Risk Level", "Findings", "Recommendations", "Extraction Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        row = 2
        for ext in self.extractions:
            ws.cell(row=row, column=1, value=ext.get("document_id", "Unknown"))
            ws.cell(row=row, column=2, value=ext.get("meta", {}).get("completeness_score", 0))
            
            risk = ext.get("categories", {}).get("risk_assessment", {}).get("risk_level", "N/A")
            ws.cell(row=row, column=3, value=risk)
            
            findings = ext.get("categories", {}).get("finding_summary", {})
            ws.cell(row=row, column=4, value="Yes" if findings.get("main_finding") else "No")
            
            recs = ext.get("categories", {}).get("recommendation_tracking", {})
            ws.cell(row=row, column=5, value="Yes" if recs.get("recommendation_text") else "No")
            
            ws.cell(row=row, column=6, value=ext.get("extraction_timestamp", "N/A"))
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_risk_sheet(self, wb):
        """Create risk assessment sheet"""
        ws = wb.create_sheet("Risk Assessment")
        
        headers = ["Document ID", "Risk Level", "Risk Description", "Risk Categories", "Mitigation Measures", "Residual Risk"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        risks = self._extract_all_risks()
        for risk in risks:
            ws.cell(row=row, column=1, value=risk.get("document"))
            ws.cell(row=row, column=2, value=risk.get("risk_level"))
            ws.cell(row=row, column=3, value=risk.get("description"))
            ws.cell(row=row, column=4, value=risk.get("categories"))
            ws.cell(row=row, column=5, value=risk.get("mitigation"))
            ws.cell(row=row, column=6, value=risk.get("residual_risk"))
            row += 1
    
    def _create_recommendations_sheet(self, wb):
        """Create recommendations sheet"""
        ws = wb.create_sheet("Recommendations")
        
        headers = ["Document ID", "Rec ID", "Recommendation", "Status", "Responsible", "Deadline"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        recs = self._extract_all_recommendations()
        for rec in recs:
            ws.cell(row=row, column=1, value=rec.get("document"))
            ws.cell(row=row, column=2, value=rec.get("rec_id"))
            ws.cell(row=row, column=3, value=rec.get("recommendation"))
            ws.cell(row=row, column=4, value=rec.get("status"))
            ws.cell(row=row, column=5, value=rec.get("responsible"))
            ws.cell(row=row, column=6, value=rec.get("deadline"))
            row += 1
    
    def _create_findings_sheet(self, wb):
        """Create findings sheet"""
        ws = wb.create_sheet("Findings")
        
        headers = ["Document ID", "Finding ID", "Main Finding", "Priority", "Affected Areas"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        findings = self._extract_all_findings()
        for finding in findings:
            ws.cell(row=row, column=1, value=finding.get("document"))
            ws.cell(row=row, column=2, value=finding.get("finding_id"))
            ws.cell(row=row, column=3, value=finding.get("finding"))
            ws.cell(row=row, column=4, value=finding.get("priority"))
            ws.cell(row=row, column=5, value=finding.get("areas"))
            row += 1
    
    def _create_metadata_sheet(self, wb):
        """Create metadata sheet"""
        ws = wb.create_sheet("Metadata")
        
        headers = ["Document ID", "Audit Type", "Audit Period", "Auditor", "Department", "Extraction Method"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="7C7C7C", end_color="7C7C7C", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 2
        metadata = self._extract_metadata()
        for meta in metadata:
            ws.cell(row=row, column=1, value=meta.get("document"))
            ws.cell(row=row, column=2, value=meta.get("audit_type"))
            ws.cell(row=row, column=3, value=meta.get("audit_period"))
            ws.cell(row=row, column=4, value=meta.get("auditor"))
            ws.cell(row=row, column=5, value=meta.get("department"))
            ws.cell(row=row, column=6, value=meta.get("extraction_method"))
            row += 1
    
    def _extract_all_risks(self) -> List[Dict[str, Any]]:
        """Extract all risk data"""
        risks = []
        for ext in self.extractions:
            risk_data = ext.get("categories", {}).get("risk_assessment", {})
            if risk_data:
                risks.append({
                    "document": ext.get("document_id", "Unknown"),
                    "risk_level": risk_data.get("risk_level", "unknown"),
                    "description": risk_data.get("risk_description", ""),
                    "categories": ", ".join(risk_data.get("risk_categories", [])),
                    "mitigation": ", ".join(risk_data.get("mitigation_measures", [])),
                    "residual_risk": risk_data.get("residual_risk", "N/A")
                })
        return risks
    
    def _extract_all_recommendations(self) -> List[Dict[str, Any]]:
        """Extract all recommendations"""
        recs = []
        for ext in self.extractions:
            rec_data = ext.get("categories", {}).get("recommendation_tracking", {})
            if rec_data.get("recommendation_text"):
                recs.append({
                    "document": ext.get("document_id", "Unknown"),
                    "rec_id": rec_data.get("recommendation_id", "N/A"),
                    "recommendation": rec_data.get("recommendation_text", ""),
                    "status": rec_data.get("status", "unknown"),
                    "responsible": rec_data.get("responsible_entity", "N/A"),
                    "deadline": rec_data.get("deadline", "N/A")
                })
        return recs
    
    def _extract_all_findings(self) -> List[Dict[str, Any]]:
        """Extract all findings"""
        findings = []
        for ext in self.extractions:
            finding_data = ext.get("categories", {}).get("finding_summary", {})
            if finding_data.get("main_finding"):
                findings.append({
                    "document": ext.get("document_id", "Unknown"),
                    "finding_id": finding_data.get("finding_id", "N/A"),
                    "finding": finding_data.get("main_finding", ""),
                    "priority": finding_data.get("priority", "unknown"),
                    "areas": ", ".join(finding_data.get("affected_areas", []))
                })
        return findings
    
    def _extract_metadata(self) -> List[Dict[str, Any]]:
        """Extract metadata"""
        metadata = []
        for ext in self.extractions:
            meta_data = ext.get("categories", {}).get("audit_metadata", {})
            metadata.append({
                "document": ext.get("document_id", "Unknown"),
                "audit_type": meta_data.get("audit_type", "N/A"),
                "audit_period": meta_data.get("audit_period", "N/A"),
                "auditor": meta_data.get("auditor", "N/A"),
                "department": meta_data.get("department", "N/A"),
                "extraction_method": ext.get("extraction_method", "N/A")
            })
        return metadata


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(description="Export MetaSynthesizer results")
    parser.add_argument(
        "--format",
        choices=["excel", "csv", "summary", "all"],
        default="all",
        help="Export format"
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("exports"),
        help="Output directory or file"
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=Path("03_extracted_data"),
        help="Directory containing extraction results"
    )
    
    args = parser.parse_args()
    
    # Create output directory if needed
    if args.format in ["csv", "all"]:
        args.output.mkdir(exist_ok=True)
    
    # Initialize exporter
    exporter = ExportManager(data_dir=args.data_dir)
    
    # Perform exports
    if args.format == "excel" or args.format == "all":
        excel_path = args.output / "metasynthesizer_results.xlsx" if args.output.is_dir() else args.output
        exporter.export_to_excel(excel_path)
    
    if args.format == "csv" or args.format == "all":
        csv_dir = args.output / "csv" if args.format == "all" else args.output
        exporter.export_to_csv(csv_dir)
    
    if args.format == "summary" or args.format == "all":
        summary_path = args.output / "extraction_summary.md" if args.output.is_dir() else args.output
        exporter.export_summary_report(summary_path)
    
    print("\n✅ All exports completed successfully!")


if __name__ == "__main__":
    main()
